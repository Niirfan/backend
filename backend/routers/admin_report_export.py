"""
backend/routers/report_export.py
ระบบ export รายงานสถิติ → PDF / Excel (แนวตั้งทั้งหมด พร้อมสไตล์มินิมอลสุภาพ)
ใช้ font และ style จาก backend/utils/pdf_styles.py
"""

from io import BytesIO
from datetime import datetime
import os
from typing import Optional
from sqlalchemy import inspect, text
from sqlalchemy.exc import SQLAlchemyError
from inspect import signature
from xml.sax.saxutils import escape

from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError
# ── ReportLab ──────────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm, mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.platypus import Image as RLImage  # เปลียนนามแฝงป้องกัน NameError ชนกับตัวอื่น

# ── openpyxl ───────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── FastAPI ────────────────────────────────────────────────────────────────
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

# ── Reuse existing font/style from pdf_styles ──────────────────────────────
from backend.utils.pdf_styles import (
    th as _th,
    THAI_FONT      as _FONT,
    THAI_FONT_BOLD as _FONT_BOLD,
    ORG_NAME_TH,
    ORG_NAME_EN,
)

# ── Internal ───────────────────────────────────────────────────────────────
from backend.database import get_db
from backend.login.dependencies import verify_admin

_BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
_LOGO_PATH = os.path.join(_BASE_DIR, "..", "static", "logo.png")

# ══════════════════════════════════════════════════════════════════════════
#  COLORS (แนวทางการ เรียบหรู ไม่เด่นแย่งซีนโลโก้ และอ่านง่าย)
# ══════════════════════════════════════════════════════════════════════════
# สำหรับ ReportLab (PDF)
_C_NAVY   = colors.HexColor("#f1f5f9")  # สีเทาอ่อนมาก (Slate 100) สำหรับพื้นหลังหัวตาราง
_C_TEXT   = colors.HexColor("#1f2937")  # สีเทาเข้ม (Gray 800) สำหรับข้อความหัวตาราง
_C_STRIPE = colors.HexColor("#f9fafb")  # สีเทาจางมาก สำหรับแถวสลับ
_C_TOTAL  = colors.HexColor("#f3f4f6")  # สีเทาอ่อน สำหรับแถวสรุปผลรวม
_C_LINE   = colors.HexColor("#374151")  # สีเทาเข้มสุภาพ สำหรับเส้นใต้ Header หลัก

# สำหรับ openpyxl (Excel)
_XL_HDR_FILL   = "F1F5F9"  
_XL_HDR_FONT   = "1F2937"  
_XL_STRIPE     = "F9FAFB"  
_XL_TOTAL_FILL = "F3F4F6"  
_XL_BORDER     = "E5E7EB"  


# ══════════════════════════════════════════════════════════════════════════
#  PDF + EXCEL HELPERS
# ══════════════════════════════════════════════════════════════════════════

_DOC_MARGIN = 14 * mm
_CONTENT_W = A4[0] - (_DOC_MARGIN * 2)


def _ps(name, font=None, size=11, bold=False, align=0, color=colors.black, space_after=4):
    return ParagraphStyle(
        name,
        fontName=(_FONT_BOLD if bold else _FONT) if font is None else font,
        fontSize=size,
        leading=size + 3,
        alignment=align,
        textColor=color,
        spaceAfter=space_after,
        wordWrap="CJK",
        splitLongWords=1,
    )


def _safe(v) -> str:
    if v is None or v == "":
        return "-"
    return escape(str(v))


def _p(v, name="cell", size=8.5, bold=False, align=0, color=colors.black):
    return Paragraph(_safe(v), _ps(name, size=size, bold=bold, align=align, color=color, space_after=0))


def _hp(html, name="html", size=8.5, bold=False, align=0, color=colors.black):
    return Paragraph(str(html), _ps(name, size=size, bold=bold, align=align, color=color, space_after=0))


def _fit_widths(widths):
    total = sum(widths)
    return [(w / total) * _CONTENT_W for w in widths]


def _new_doc(buf):
    return SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=_DOC_MARGIN,
        rightMargin=_DOC_MARGIN,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )


def _pdf_header(story, report_title: str, subtitle: str = ""):
    logo_cell = ""
    if os.path.exists(_LOGO_PATH):
        logo_cell = RLImage(_LOGO_PATH, width=17 * mm, height=17 * mm)

    org_block = [
        Paragraph(ORG_NAME_TH, _ps("hOrgTh", size=12, bold=True, space_after=1)),
        Paragraph(ORG_NAME_EN, _ps("hOrgEn", size=8, color=colors.gray, space_after=0)),
    ]

    title_block = [
        Paragraph(report_title, _ps("hTitle", size=13, bold=True, align=2, color=_C_LINE, space_after=1)),
    ]
    if subtitle:
        title_block.append(Paragraph(subtitle, _ps("hSub", size=8, align=2, color=colors.grey, space_after=0)))

    top = Table(
        [[logo_cell, org_block, title_block]],
        colWidths=[20 * mm, 74 * mm, _CONTENT_W - (94 * mm)],
        hAlign="CENTER",
    )
    top.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "LEFT"),
        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(top)

    line = Table([[""]], colWidths=[_CONTENT_W], hAlign="CENTER")
    line.setStyle(TableStyle([
        ("LINEBELOW", (0, 0), (-1, -1), 0.8, _C_LINE),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(line)
    story.append(Spacer(1, 5 * mm))


def _table_style(has_total=False, total_row=0):
    s = [
        ("BACKGROUND", (0, 0), (-1, 0), _C_NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), _C_TEXT),
        ("FONTNAME", (0, 0), (-1, 0), _FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, 0), 8.5),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d1d5db")),
        ("FONTNAME", (0, 1), (-1, -1), _FONT),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]

    if has_total and total_row > 1:
        s.append(("ROWBACKGROUNDS", (0, 1), (-1, total_row - 1), [colors.white, _C_STRIPE]))
    elif not has_total:
        s.append(("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _C_STRIPE]))

    if has_total and total_row > 0:
        s += [
            ("BACKGROUND", (0, total_row), (-1, total_row), _C_TOTAL),
            ("FONTNAME", (0, total_row), (-1, total_row), _FONT_BOLD),
            ("TEXTCOLOR", (0, total_row), (-1, total_row), colors.black),
        ]

    return TableStyle(s)


def _pdf_table(rows, col_widths, has_total=False, total_row=0):
    t = Table(
        rows,
        colWidths=_fit_widths(col_widths),
        repeatRows=1,
        hAlign="CENTER",
        splitByRow=1,
    )
    t.setStyle(_table_style(has_total=has_total, total_row=total_row))
    return t


def _summary_box(text: str):
    box = Table(
        [[Paragraph(_safe(text), _ps("summary", size=10.5, bold=True, align=1, color=colors.HexColor("#1f2937")))]],
        colWidths=[_CONTENT_W],
        hAlign="CENTER",
    )
    box.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor("#d1d5db")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    return box

def _pdf_footer_timestamp(story):
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(
        f"พิมพ์เมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        _ps("footer", size=7.5, align=2, color=colors.grey, space_after=0),
    ))


def _border():
    s = Side(style="thin", color=_XL_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _xl_title_rows(ws, title: str, col_count: int):
    ws.insert_rows(1, 3)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_count)

    t = ws.cell(row=1, column=1, value=ORG_NAME_TH)
    t.font = Font(name="Arial", bold=True, size=13, color="374151")
    t.alignment = Alignment(horizontal="center")

    s = ws.cell(row=2, column=1, value=title)
    s.font = Font(name="Arial", bold=True, size=11, color="374151")
    s.alignment = Alignment(horizontal="center", wrap_text=True)

    d = ws.cell(row=3, column=1, value=f"วันที่พิมพ์: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    d.font = Font(name="Arial", size=9, color="888888")
    d.alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 14


def _xl_header(ws, headers: list[str], row: int):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color=_XL_HDR_FONT, size=10)
        cell.fill = PatternFill("solid", fgColor=_XL_HDR_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[row].height = 22


def _xl_row(ws, values: list, row: int, is_total=False, currency_cols: set = None):
    bg = _XL_TOTAL_FILL if is_total else (_XL_STRIPE if row % 2 == 0 else "FFFFFF")
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = Font(name="Arial", bold=is_total, size=9)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = _border()
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        if currency_cols and c in currency_cols:
            cell.number_format = "#,##0.00"
        elif isinstance(v, (int, float)) and not isinstance(v, bool):
            cell.number_format = "#,##0"


def _xl_autowidth(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[letter].width = min(width + 4, 42)


def _xl_freeze(ws, row=8):
    ws.freeze_panes = ws.cell(row=row, column=1)


def _branch_label(db: Session, branch_id: Optional[str], payload=None) -> str:
    if not branch_id:
        return "ทุกสาขา"

    def pick(obj):
        if isinstance(obj, dict):
            value = obj.get("branch_name") or obj.get("branch_label")
            if isinstance(value, str) and value.strip():
                return value.strip()

            branch = obj.get("branch")
            if isinstance(branch, dict):
                value = branch.get("branch_name")
                if isinstance(value, str) and value.strip():
                    return value.strip()

            for key in ("items", "requests", "data", "rows"):
                label = pick(obj.get(key))
                if label:
                    return label

        if isinstance(obj, list):
            for item in obj:
                label = pick(item)
                if label:
                    return label

        return None

    label = pick(payload)
    if label:
        return label

    try:
        row = db.execute(
            text("""
                SELECT branch_name
                FROM sams_branch
                WHERE branch_id = :branch_id
                LIMIT 1
            """),
            {"branch_id": str(branch_id)},
        ).mappings().first()

        if row and row.get("branch_name"):
            return str(row["branch_name"]).strip()

    except SQLAlchemyError:
        db.rollback()

    return f"สาขา {branch_id}"


# ══════════════════════════════════════════════════════════════════════════
#  1. MONTHLY
# ══════════════════════════════════════════════════════════════════════════

def export_monthly_pdf(year: int, data: list[dict], branch_label: str = "ทุกสาขา") -> BytesIO:
    buf = BytesIO()
    doc = _new_doc(buf)
    story: list = []

    _pdf_header(
        story,
        f"รายงานสถิติการเบิกวัสดุรายเดือน  ปี พ.ศ. {year + 543}",
        f"สาขา: {branch_label}",
    )

    headers = ["เดือน", "จำนวนใบเบิก (ครั้ง)", "มูลค่ารวม (บาท)"]
    rows = [[_p(h, f"mh{i}", bold=True, align=1, color=_C_TEXT) for i, h in enumerate(headers)]]

    total_count = 0
    total_price = 0

    for d in data:
        count = d.get("count", 0) or 0
        total = d.get("total_price", 0) or 0
        rows.append([
            _p(d.get("month"), "mc"),
            _p(f"{count:,}", "mn", align=2),
            _p(f"{total:,.2f}", "mp", align=2),
        ])
        total_count += count
        total_price += total

    rows.append([
        _p("รวมทั้งปี", "mt", bold=True),
        _p(f"{total_count:,}", "mtc", bold=True, align=2),
        _p(f"{total_price:,.2f}", "mtp", bold=True, align=2),
    ])

    story.append(_pdf_table(rows, [38, 34, 38], has_total=True, total_row=len(rows) - 1))
    _pdf_footer_timestamp(story)

    doc.build(story)
    buf.seek(0)
    return buf


def export_monthly_excel(year: int, data: list[dict], branch_label: str = "ทุกสาขา") -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "รายเดือน"

    headers = ["เดือน", "จำนวนใบเบิก (ครั้ง)", "มูลค่ารวม (บาท)"]
    DATA_ROW = 5

    _xl_header(ws, headers, row=4)
    for i, d in enumerate(data, DATA_ROW):
        _xl_row(ws, [d.get("month"), d.get("count", 0), d.get("total_price", 0)], i, currency_cols={3})

    total_row = DATA_ROW + len(data)
    _xl_row(
        ws,
        ["รวมทั้งปี", f"=SUM(B{DATA_ROW}:B{total_row-1})", f"=SUM(C{DATA_ROW}:C{total_row-1})"],
        total_row,
        is_total=True,
        currency_cols={3},
    )

    _xl_title_rows(ws, f"รายงานสถิติการเบิกวัสดุรายเดือน ปี พ.ศ. {year + 543} | สาขา: {branch_label}", 3)
    _xl_autowidth(ws)
    _xl_freeze(ws, row=8)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════
#  1.1 MONTHLY DETAIL
# ══════════════════════════════════════════════════════════════════════════

def export_monthly_detail_pdf(year: int, month: int, data: dict, branch_label: str = "ทุกสาขา") -> BytesIO:
    from backend.config import THAI_MONTHS

    buf = BytesIO()
    doc = _new_doc(buf)
    story: list = []

    month_label = THAI_MONTHS[month]
    y_label = year + 543
    total = data.get("total", 0) or 0
    requests = data.get("requests", [])

    _pdf_header(
        story,
        "รายงานรายละเอียดการเบิกวัสดุ",
        f"เดือน{month_label} ปี พ.ศ. {y_label} | สาขา: {branch_label}",
    )

    story.append(_summary_box(f"จำนวนใบเบิกทั้งหมด: {total:,} ใบ"))
    story.append(Spacer(1, 5 * mm))

    item_headers = ["รหัสวัสดุ", "ชื่อวัสดุ", "หน่วย", "จำนวนขอ", "จำนวนอนุมัติ"]
    item_widths = [22, 78, 18, 24, 26]

    for idx, req in enumerate(requests, 1):
        raw_date = req.get("req_date") or ""
        req_date = str(raw_date)[:10] if raw_date else "-"
        try:
            dt = datetime.fromisoformat(str(raw_date))
            req_date = f"{dt.day:02d}/{dt.month:02d}/{dt.year + 543}"
        except Exception:
            pass

        req_branch = req.get("branch_name") or branch_label

        req_info = Table(
            [[
                _hp(
                    f"<b>ใบเบิกที่ {idx}/{total}</b><br/>"
                    f"เลขที่: <b>{_safe(req.get('mat_req_code'))}</b><br/>"
                    f"วันที่: <b>{_safe(req_date)}</b>",
                    "rh",
                    size=8.7,
                ),
                _hp(
                    f"รหัสพนักงาน: <b>{_safe(req.get('emp_code'))}</b><br/>"
                    f"ชื่อ: <b>{_safe(req.get('full_name'))}</b><br/>"
                    f"สาขา: <b>{_safe(req_branch)}</b>",
                    "rh2",
                    size=8.7,
                    align=2,
                ),
            ]],
            colWidths=[_CONTENT_W * 0.48, _CONTENT_W * 0.52],
            hAlign="CENTER",
        )
        req_info.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#e2e8f0")),
            ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(req_info)

        items = req.get("items", [])
        if items:
            rows = [[_p(h, f"mdh{i}", bold=True, align=1, color=_C_TEXT) for i, h in enumerate(item_headers)]]
            for it in items:
                rows.append([
                    _p(it.get("mat_code"), "id0", align=1),
                    _p(it.get("mat_name"), "id1"),
                    _p(it.get("unit"), "id2", align=1),
                    _p(f"{it.get('req_qty', 0):,}", "id3", align=2),
                    _p(f"{it.get('approve_qty', 0):,}", "id4", align=2),
                ])
            story.append(_pdf_table(rows, item_widths))
        else:
            story.append(Paragraph("- ไม่มีรายการวัสดุ -", _ps("noitem", size=8.5, align=1, color=colors.grey)))

        story.append(Spacer(1, 5 * mm))

    _pdf_footer_timestamp(story)

    doc.build(story)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════
#  2. INVENTORY VALUE
# ══════════════════════════════════════════════════════════════════════════

def export_inventory_pdf(grand_total: float, items: list[dict], branch_label: str = "ทุกสาขา") -> BytesIO:
    buf = BytesIO()
    doc = _new_doc(buf)
    story: list = []

    _pdf_header(story, "รายงานมูลค่าวัสดุคงคลัง", f"สาขา: {branch_label}")

    story.append(_summary_box(f"มูลค่าคงคลังทั้งหมด: {grand_total:,.2f} บาท"))
    story.append(Spacer(1, 4 * mm))

    headers = ["รหัสวัสดุ", "ชื่อวัสดุ", "ประเภท", "หน่วย", "คงเหลือ", "ราคา/หน่วย", "มูลค่ารวม"]
    rows = [[_p(h, f"ih{i}", bold=True, align=1, color=_C_TEXT) for i, h in enumerate(headers)]]

    for it in items:
        rows.append([
            _p(it.get("mat_code"), "ic1", align=1),
            _p(it.get("mat_name"), "ic2"),
            _p(it.get("mat_type"), "ic3", align=1),
            _p(it.get("unit"), "ic4", align=1),
            _p(f"{it.get('quantity', 0):,}", "ic5", align=2),
            _p(f"{it.get('unit_price', 0):,.2f}", "ic6", align=2),
            _p(f"{it.get('total_value', 0):,.2f}", "ic7", align=2),
        ])

    rows.append([
        _p("", "it0"),
        _p("", "it1"),
        _p("", "it2"),
        _p("", "it3"),
        _p("", "it4"),
        _p("รวมทั้งหมด", "itl", bold=True, align=2),
        _p(f"{grand_total:,.2f}", "itv", bold=True, align=2),
    ])

    story.append(_pdf_table(rows, [22, 58, 28, 18, 20, 28, 32], has_total=True, total_row=len(rows) - 1))
    _pdf_footer_timestamp(story)

    doc.build(story)
    buf.seek(0)
    return buf


def export_inventory_excel(grand_total: float, items: list[dict], branch_label: str = "ทุกสาขา") -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "คงคลัง"

    headers = ["รหัสวัสดุ", "ชื่อวัสดุ", "ประเภท", "หน่วย", "คงเหลือ", "ราคา/หน่วย", "มูลค่ารวม"]
    DATA_ROW = 5

    _xl_header(ws, headers, row=4)
    for i, it in enumerate(items, DATA_ROW):
        _xl_row(
            ws,
            [
                it.get("mat_code"),
                it.get("mat_name"),
                it.get("mat_type"),
                it.get("unit"),
                it.get("quantity", 0),
                it.get("unit_price", 0),
                it.get("total_value", 0),
            ],
            i,
            currency_cols={6, 7},
        )

    total_row = DATA_ROW + len(items)
    _xl_row(
        ws,
        ["", "", "", "", "", "มูลค่าทั้งหมด", f"=SUM(G{DATA_ROW}:G{total_row-1})"],
        total_row,
        is_total=True,
        currency_cols={7},
    )

    _xl_title_rows(ws, f"รายงานมูลค่าวัสดุคงคลัง | สาขา: {branch_label}", 7)
    _xl_autowidth(ws)
    _xl_freeze(ws, row=8)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════
#  3. TOP MATERIALS
# ══════════════════════════════════════════════════════════════════════════

def export_top_materials_pdf(year: int | None, items: list[dict], branch_label: str = "ทุกสาขา") -> BytesIO:
    y_label = f"ปี พ.ศ. {year + 543}" if year else "ทุกปี"
    buf = BytesIO()
    doc = _new_doc(buf)
    story: list = []

    _pdf_header(story, f"รายงานวัสดุที่เบิกมากที่สุด ({y_label})", f"สาขา: {branch_label}")

    headers = ["#", "รหัสวัสดุ", "ชื่อวัสดุ", "ประเภท", "หน่วย", "จำนวนเบิก", "จำนวนครั้ง"]
    rows = [[_p(h, f"th{i}", bold=True, align=1, color=_C_TEXT) for i, h in enumerate(headers)]]

    for rank, it in enumerate(items, 1):
        rows.append([
            _p(rank, "tr0", align=1),
            _p(it.get("mat_code"), "tr1", align=1),
            _p(it.get("mat_name"), "tr2"),
            _p(it.get("mat_type"), "tr3", align=1),
            _p(it.get("unit"), "tr4", align=1),
            _p(f"{it.get('total_qty', 0):,}", "tr5", align=2),
            _p(f"{it.get('req_count', 0):,}", "tr6", align=2),
        ])

    story.append(_pdf_table(rows, [10, 24, 58, 28, 18, 24, 22]))
    _pdf_footer_timestamp(story)

    doc.build(story)
    buf.seek(0)
    return buf


def export_top_materials_excel(year: int | None, items: list[dict], branch_label: str = "ทุกสาขา") -> BytesIO:
    y_label = f"ปี พ.ศ. {year + 543}" if year else "ทุกปี"
    wb = Workbook()
    ws = wb.active
    ws.title = "วัสดุยอดนิยม"

    headers = ["อันดับ", "รหัสวัสดุ", "ชื่อวัสดุ", "ประเภท", "หน่วย", "จำนวนเบิก", "จำนวนครั้ง"]
    DATA_ROW = 5

    _xl_header(ws, headers, row=4)
    for i, it in enumerate(items, DATA_ROW):
        _xl_row(
            ws,
            [
                i - DATA_ROW + 1,
                it.get("mat_code"),
                it.get("mat_name"),
                it.get("mat_type"),
                it.get("unit"),
                it.get("total_qty", 0),
                it.get("req_count", 0),
            ],
            i,
        )

    _xl_title_rows(ws, f"รายงานวัสดุที่เบิกมากที่สุด ({y_label}) | สาขา: {branch_label}", 7)
    _xl_autowidth(ws)
    _xl_freeze(ws, row=8)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════
#  4. BY-USER
# ══════════════════════════════════════════════════════════════════════════

def export_by_user_pdf(year: int | None, items: list[dict], branch_label: str = "ทุกสาขา") -> BytesIO:
    y_label = f"ปี พ.ศ. {year + 543}" if year else "ทุกปี"
    buf = BytesIO()
    doc = _new_doc(buf)
    story: list = []

    _pdf_header(story, f"รายงานสถิติการเบิกรายบุคคล ({y_label})", f"สาขา: {branch_label}")

    headers = ["รหัสพนักงาน", "ชื่อ-สกุล", "สาขา / ส่วนงาน", "จำนวนใบเบิก", "มูลค่ารวม"]
    rows = [[_p(h, f"uh{i}", bold=True, align=1, color=_C_TEXT) for i, h in enumerate(headers)]]

    total_count = 0
    total_price = 0

    for it in items:
        req_count = it.get("req_count", 0) or 0
        total = it.get("total_price", 0) or 0

        rows.append([
            _p(it.get("emp_code"), "ur0", align=1),
            _p(it.get("full_name"), "ur1"),
            _p(it.get("branch_name") or "-", "ur2"),
            _p(f"{req_count:,}", "ur3", align=2),
            _p(f"{total:,.2f}", "ur4", align=2),
        ])

        total_count += req_count
        total_price += total

    rows.append([
        _p("", "ut0"),
        _p("รวมทั้งหมด", "ut1", bold=True),
        _p("", "ut2"),
        _p(f"{total_count:,}", "ut3", bold=True, align=2),
        _p(f"{total_price:,.2f}", "ut4", bold=True, align=2),
    ])

    story.append(_pdf_table(rows, [28, 56, 44, 24, 30], has_total=True, total_row=len(rows) - 1))
    _pdf_footer_timestamp(story)

    doc.build(story)
    buf.seek(0)
    return buf


def export_by_user_excel(year: int | None, items: list[dict], branch_label: str = "ทุกสาขา") -> BytesIO:
    y_label = f"ปี พ.ศ. {year + 543}" if year else "ทุกปี"
    wb = Workbook()
    ws = wb.active
    ws.title = "รายบุคคล"

    headers = ["รหัสพนักงาน", "ชื่อ-สกุล", "สาขา / ส่วนงาน", "จำนวนใบเบิก", "มูลค่ารวม"]
    DATA_ROW = 5

    _xl_header(ws, headers, row=4)
    for i, it in enumerate(items, DATA_ROW):
        _xl_row(
            ws,
            [
                it.get("emp_code"),
                it.get("full_name"),
                it.get("branch_name") or "-",
                it.get("req_count", 0),
                it.get("total_price", 0),
            ],
            i,
            currency_cols={5},
        )

    total_row = DATA_ROW + len(items)
    _xl_row(
        ws,
        ["", "รวมทั้งหมด", "", f"=SUM(D{DATA_ROW}:D{total_row-1})", f"=SUM(E{DATA_ROW}:E{total_row-1})"],
        total_row,
        is_total=True,
        currency_cols={5},
    )

    _xl_title_rows(ws, f"รายงานสถิติการเบิกรายบุคคล ({y_label}) | สาขา: {branch_label}", 5)
    _xl_autowidth(ws)
    _xl_freeze(ws, row=8)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════
#  FASTAPI ROUTER
# ══════════════════════════════════════════════════════════════════════════

router = APIRouter(prefix="/admin/report", tags=["Admin Report Export"])

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME = "application/pdf"


def _stream(buf: BytesIO, media: str, filename: str) -> StreamingResponse:
    return StreamingResponse(
        buf,
        media_type=media,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/monthly/export")
def export_monthly(
    year: Optional[int] = Query(default=None),
    format: str = Query(default="pdf", pattern="^(pdf|excel)$"),
    branch_id: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(verify_admin),
):
    from backend.routers.admin_report import report_monthly

    result = report_monthly(year=year, branch_id=branch_id, db=db, admin_user=_)
    y, data = result["year"], result["items"]
    branch_label = _branch_label(db, branch_id, result)

    suffix = f"_{branch_id}" if branch_id else ""

    if format == "excel":
        return _stream(export_monthly_excel(y, data, branch_label), EXCEL_MIME, f"monthly_{y}{suffix}.xlsx")

    return _stream(export_monthly_pdf(y, data, branch_label), PDF_MIME, f"monthly_{y}{suffix}.pdf")


@router.get("/monthly-detail/export")
def export_monthly_detail(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    branch_id: Optional[str] = Query(default=None),
    format: str = Query(default="pdf", pattern="^(pdf|excel)$"),
    db: Session = Depends(get_db),
    _: dict = Depends(verify_admin),
):
    from backend.routers.admin_report import report_monthly_detail

    data = report_monthly_detail(year=year, month=month, branch_id=branch_id, db=db, admin_user=_)
    branch_label = _branch_label(db, branch_id, data)

    suffix = f"_{branch_id}" if branch_id else ""

    if format == "pdf":
        buf = export_monthly_detail_pdf(year, month, data, branch_label)
        return _stream(buf, PDF_MIME, f"monthly_detail_{year}_{month:02d}{suffix}.pdf")

    return {"detail": "Excel format not yet supported for this report"}


@router.get("/inventory-value/export")
def export_inventory(
    format: str = Query(default="pdf", pattern="^(pdf|excel)$"),
    branch_id: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(verify_admin),
):
    from backend.routers.admin_report import report_inventory_value

    kwargs = {"db": db, "admin_user": _}
    if "branch_id" in signature(report_inventory_value).parameters:
        kwargs["branch_id"] = branch_id

    result = report_inventory_value(**kwargs)
    branch_label = _branch_label(db, branch_id, result)

    suffix = f"_{branch_id}" if branch_id else ""

    if format == "excel":
        return _stream(
            export_inventory_excel(result["grand_total"], result["items"], branch_label),
            EXCEL_MIME,
            f"inventory_value{suffix}.xlsx",
        )

    return _stream(
        export_inventory_pdf(result["grand_total"], result["items"], branch_label),
        PDF_MIME,
        f"inventory_value{suffix}.pdf",
    )


@router.get("/top-materials/export")
def export_top_mats(
    limit: int = Query(default=10, ge=1, le=50),
    year: Optional[int] = Query(default=None),
    format: str = Query(default="pdf", pattern="^(pdf|excel)$"),
    branch_id: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(verify_admin),
):
    from backend.routers.admin_report import report_top_materials

    items = report_top_materials(limit=limit, year=year, branch_id=branch_id, db=db, admin_user=_)
    branch_label = _branch_label(db, branch_id, items)

    suffix = f"_{year}" if year else ""
    if branch_id:
        suffix += f"_{branch_id}"

    if format == "excel":
        return _stream(export_top_materials_excel(year, items, branch_label), EXCEL_MIME, f"top_materials{suffix}.xlsx")

    return _stream(export_top_materials_pdf(year, items, branch_label), PDF_MIME, f"top_materials{suffix}.pdf")


@router.get("/by-user/export")
def export_by_user_route(
    year: Optional[int] = Query(default=None),
    format: str = Query(default="pdf", pattern="^(pdf|excel)$"),
    branch_id: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(verify_admin),
):
    from backend.routers.admin_report import report_by_user

    items = report_by_user(year=year, branch_id=branch_id, db=db, admin_user=_)
    branch_label = _branch_label(db, branch_id, items)

    suffix = f"_{year}" if year else ""
    if branch_id:
        suffix += f"_{branch_id}"

    if format == "excel":
        return _stream(export_by_user_excel(year, items, branch_label), EXCEL_MIME, f"by_user{suffix}.xlsx")

    return _stream(export_by_user_pdf(year, items, branch_label), PDF_MIME, f"by_user{suffix}.pdf")