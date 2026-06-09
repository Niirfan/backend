# backend/routers/inventory.py
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from io import BytesIO
import io
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import ParagraphStyle

from backend.database import get_db
from backend.models.material import Material, MaterialStock
from backend.models.request import MaterialReserved, ReservedStatus
from backend.models.master import MaterialType
from backend.login.dependencies import verify_admin
from backend.utils.pdf_styles import th as _th, THAI_FONT, ORG_NAME_TH, ORG_NAME_EN

router = APIRouter(prefix="/inventory", tags=["Inventory"])

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_LOGO_PATH = os.path.join(_BASE_DIR, "..", "static", "logo.png")


def fetch_inventory_data(db: Session):
    stock_sub = (
        db.query(
            MaterialStock.mat_id,
            func.coalesce(func.sum(MaterialStock.quantity), 0).label("total_stock"),
            func.max(MaterialStock.import_date).label("last_import")
        )
        .group_by(MaterialStock.mat_id)
        .subquery()
    )

    reserved_sub = (
        db.query(
            MaterialReserved.mat_id,
            func.coalesce(func.sum(MaterialReserved.quantity), 0).label("total_reserved")
        )
        .filter(MaterialReserved.status.in_([ReservedStatus.RESERVED, ReservedStatus.APPROVED]))
        .group_by(MaterialReserved.mat_id)
        .subquery()
    )

    results = (
        db.query(
            Material.mat_id,
            Material.mat_code,
            Material.mat_name,
            MaterialType.mat_type_name,
            Material.unit_pack,
            stock_sub.c.total_stock,
            stock_sub.c.last_import,
            func.coalesce(reserved_sub.c.total_reserved, 0).label("total_reserved"),
        )
        .join(MaterialType, Material.mat_type_id == MaterialType.mat_type_id)
        .outerjoin(stock_sub, Material.mat_id == stock_sub.c.mat_id)
        .outerjoin(reserved_sub, Material.mat_id == reserved_sub.c.mat_id)
        .filter(Material.is_active == True)
        .order_by(Material.mat_code.asc())
        .all()
    )

    rows = []
    for r in results:
        stock = r.total_stock or 0
        reserved = r.total_reserved or 0
        rows.append({
            "id": r.mat_code,
            "name": r.mat_name,
            "type": r.mat_type_name,
            "stock": stock,
            "unit": r.unit_pack,
            "date": r.last_import.strftime("%d/%m/%Y") if r.last_import else "-"
        })
    return rows


@router.get("")
def get_inventory(
    db: Session = Depends(get_db),
    admin_user: dict = Depends(verify_admin)
):
    return fetch_inventory_data(db)


@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    admin_user: dict = Depends(verify_admin)
):
    rows = fetch_inventory_data(db)
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายงานวัสดุคงคลัง"

    header_fill = PatternFill("solid", fgColor="157282")
    header_font = Font(bold=True, color="FFFFFF", size=12, name="TH Sarabun New")
    data_font   = Font(size=12, name="TH Sarabun New")
    title_font  = Font(bold=True, size=16, color="157282", name="TH Sarabun New")
    sub_font    = Font(size=11, color="888888", name="TH Sarabun New")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill    = PatternFill("solid", fgColor="F0F9FA")

    ws.merge_cells("A1:G1")
    ws["A1"].value     = f"รายงานวัสดุคงคลัง — {ORG_NAME_TH}"
    ws["A1"].font      = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    ws["A2"].value     = f"พิมพ์เมื่อ: {now_str}"
    ws["A2"].font      = sub_font
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 20

    headers    = ["ลำดับ", "รหัสวัสดุ", "ชื่อวัสดุ", "ประเภท", "คงเหลือ", "หน่วย", "วันที่นำเข้าล่าสุด"]
    col_widths = [8, 15, 38, 22, 12, 12, 22]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[3].height = 28

    for row_idx, item in enumerate(rows, start=4):
        values = [row_idx - 3, item["id"], item["name"], item["type"], item["stock"], item["unit"], item["date"]]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = data_font
            cell.alignment = left if col == 3 else center
            cell.border    = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory.xlsx"}
    )


@router.get("/export/pdf")
def export_pdf(
    db: Session = Depends(get_db),
    admin_user: dict = Depends(verify_admin)
):
    rows = fetch_inventory_data(db)
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )

    story = []

    logo_cell = ""
    if os.path.exists(_LOGO_PATH):
        logo_cell = Image(_LOGO_PATH, width=22*mm, height=22*mm)

    org_block = [
        Paragraph(ORG_NAME_TH, _th(14, bold=True)),
        Paragraph(ORG_NAME_EN, _th(10)),
    ]
    doc_title_block = [
        Paragraph("รายงานวัสดุคงคลัง", _th(16, bold=True, align=2)),
        Paragraph("Inventory Report",   _th(10, align=2)),
    ]

    header_top = Table(
        [[logo_cell, org_block, doc_title_block]],
        colWidths=[28*mm, 70*mm, 72*mm]
    )
    header_top.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "CENTER"),
        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
    ]))
    story.append(header_top)

    story.append(Spacer(1, 2*mm))
    story.append(Table([[""]], colWidths=[170*mm], style=TableStyle([
        ("LINEBELOW", (0, 0), (-1, -1), 1.0, colors.black),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ])))
    story.append(Spacer(1, 5*mm))

    info_table = Table([
        [
            Paragraph(f"<b>วันที่พิมพ์:</b> {now_str}", _th(11)),
            Paragraph(f"<b>จำนวนรายการ:</b> {len(rows)} รายการ", _th(11)),
        ],
    ], colWidths=[85*mm, 85*mm])
    info_table.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("BOX",           (0, 0), (-1, -1), 0.5, colors.black),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 8*mm))

    table_data = [[
        Paragraph("ลำดับ",              _th(10, True, align=1)),
        Paragraph("รหัสวัสดุ",          _th(10, True, align=1)),
        Paragraph("ชื่อวัสดุ",           _th(10, True, align=1)),
        Paragraph("ประเภท",             _th(10, True, align=1)),
        Paragraph("คงเหลือ",            _th(10, True, align=1)),
        Paragraph("หน่วย",              _th(10, True, align=1)),
        Paragraph("วันที่นำเข้าล่าสุด", _th(10, True, align=1)),
    ]]

    for i, item in enumerate(rows, 1):
        table_data.append([
            Paragraph(str(i),             _th(10, align=1)),
            Paragraph(item["id"],         _th(10, align=1)),
            Paragraph(item["name"],       _th(10)),
            Paragraph(item["type"],       _th(10)),
            Paragraph(str(item["stock"]), _th(10, align=1)),
            Paragraph(item["unit"],       _th(10, align=1)),
            Paragraph(item["date"],       _th(10, align=1)),
        ])

    item_table = Table(
        table_data,
        colWidths=[10*mm, 20*mm, 50*mm, 28*mm, 16*mm, 16*mm, 30*mm],
        repeatRows=1
    )
    item_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  colors.lightgrey),
        ("TEXTCOLOR",     (0, 0), (-1, -1), colors.black),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.black),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F9FA")]),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
    ]))
    story.append(item_table)

    story.append(Spacer(1, 10*mm))
    story.append(Paragraph(
        now_str,
        ParagraphStyle("footer", fontName=THAI_FONT, fontSize=8,
                       textColor=colors.grey, alignment=2)
    ))

    doc.build(story)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=inventory.pdf"}
    )